"""
Task completion interface for IMS maintenance checklists.

This module provides a GUI for systematically completing maintenance tasks
with user input validation and progress tracking.

TODO:
ISSUES:
- When saving to Excel, preserve existing data and formatting, as well any input cell data if it exists and the new data is not a new month.
"""

from tkinter import *
from tkinter import messagebox, ttk
from datetime import datetime
import calendar
from typing import Callable, Dict, List, Optional
from pathlib import Path
import json
import logging
import openpyxl  # Add this import
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class TaskCompletionUI:
    """Handles the UI for completing maintenance tasks."""
    
    def __init__(self, json_path: Path):
        """Initialize the task completion interface."""
        self.root = Tk()
        self.root.title("IMS Task Completion")
        self.root.geometry("600x400")
        
        self.json_path = json_path  # Store json_path for later use
        
        # Load task data
        self.data = self._load_and_validate_data(json_path)
        
        # Track current position in tasks
        self.current_sheet = 0
        self.current_category = 0
        self.current_task = 0
        
        # Store user details
        self.initials = StringVar()  # Changed from user_initials to initials
        self.selected_day = StringVar()
        self.selected_month = StringVar()
        self.selected_year = StringVar()
        
        self._setup_ui()
        
    def _setup_ui(self):
        """Create and configure the UI elements."""
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # User details frame
        details_frame = ttk.LabelFrame(main_frame, text="User Details", padding="5")
        details_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        # Initials entry
        ttk.Label(details_frame, text="Initials:").grid(row=0, column=0, padx=5)
        initials_entry = ttk.Entry(
            details_frame, 
            textvariable=self.initials,  # Changed from user_initials to initials
            width=5,
            validate="key",
            validatecommand=(self.root.register(self._validate_initials), '%P')
        )
        initials_entry.grid(row=0, column=1, padx=5)
        
        # Date selection
        ttk.Label(details_frame, text="Date:").grid(row=0, column=2, padx=5)
        
        # Day dropdown
        days = [str(i).zfill(2) for i in range(1, 32)]
        day_cb = ttk.Combobox(
            details_frame, 
            textvariable=self.selected_day,
            values=days,
            width=3,
            state="readonly"
        )
        day_cb.grid(row=0, column=3, padx=2)
        
        # Month dropdown
        months = list(calendar.month_abbr)[1:]
        month_cb = ttk.Combobox(
            details_frame, 
            textvariable=self.selected_month,
            values=months,
            width=4,
            state="readonly"
        )
        month_cb.grid(row=0, column=4, padx=2)
        
        # Year dropdown
        current_year = datetime.now().year
        years = [str(i) for i in range(current_year-1, current_year+2)]
        year_cb = ttk.Combobox(
            details_frame, 
            textvariable=self.selected_year,
            values=years,
            width=6,
            state="readonly"
        )
        year_cb.grid(row=0, column=5, padx=2)
        
        # Set defaults
        today = datetime.now()
        self.selected_day.set(today.strftime("%d"))
        self.selected_month.set(today.strftime("%b"))
        self.selected_year.set(str(today.year))
        
        # Continue button
        ttk.Button(
            details_frame,
            text="Begin Tasks",
            command=self._validate_and_start
        ).grid(row=0, column=6, padx=10)
        
    def _validate_initials(self, value: str) -> bool:
        """
        Validate initials input.
        
        Args:
            value: The string to validate
            
        Returns:
            bool: True if input is valid, False otherwise
        """
        # Allow empty string for deletion
        if not value:
            return True
            
        # Check length and characters
        if len(value) <= 3 and value.isalpha():
            return True
            
        return False
        
    def _validate_and_start(self):
        """Validate user input and start task completion."""
        # Check initials
        if len(self.initials.get()) < 2:  # Changed from user_initials to initials
            messagebox.showerror(
                "Invalid Input",
                "Please enter at least 2 characters for initials"
            )
            return
            
        # Validate date
        try:
            # Convert month abbreviation to number
            month_num = list(calendar.month_abbr).index(self.selected_month.get())
            
            # Create date object to validate
            datetime(
                int(self.selected_year.get()),
                month_num,
                int(self.selected_day.get())
            )
        except (ValueError, IndexError):
            messagebox.showerror(
                "Invalid Date",
                "Please select a valid date"
            )
            return
            
        # Start task completion
        self._show_task_interface()

      
    def _show_task_interface(self):
        """Display the task completion interface."""
        # Clear previous content
        for widget in self.root.winfo_children():
            widget.destroy()
        
        # Configure main window layout
        self.root.grid_rowconfigure(0, weight=1)  # Make content expandable
        self.root.grid_columnconfigure(0, weight=1)
        
        # Create main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Configure main frame layout
        main_frame.grid_rowconfigure(1, weight=1)  # Make task frame expandable
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Create header frame - fixed at top
        self._create_header_frame(main_frame)
        
        # Create task content frame - expandable middle
        self._create_task_frame(main_frame)
        
        # Create navigation frame - fixed at bottom right
        self._create_navigation_frame(main_frame)

    def _create_header_frame(self, parent):
        """Create header section with user details."""
        header_frame = ttk.Frame(parent)
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        # Display current user and date
        user_text = f"User: {self.initials.get().upper()}"
        date_text = (f"Date: {self.selected_day.get()}-"
                    f"{self.selected_month.get().upper()}-"
                    f"{self.selected_year.get()}")
        ttk.Label(header_frame, text=user_text).grid(row=0, column=0, padx=5)
        ttk.Label(header_frame, text=date_text).grid(row=0, column=1, padx=5)

    def _create_task_frame(self, parent):
        """Create main task content area."""
        # Get current task data
        current_task = self._get_current_task()
        if not current_task:
            return
            
        # Create task frame with padding
        task_frame = ttk.LabelFrame(parent, text="Current Task", padding="10")
        task_frame.grid(row=1, column=0, sticky="nsew", pady=5)
        task_frame.grid_columnconfigure(0, weight=1)
        
        # Display task information
        current_category = self.data[self.current_sheet]['categories'][self.current_category]
        
        # Category name with bold font
        ttk.Label(
            task_frame,
            text=f"Category: {current_category['name']}",
            font=('Helvetica', 10, 'bold'),
            wraplength=500
        ).grid(row=0, column=0, sticky="w", pady=5)
        
        # Task name
        ttk.Label(
            task_frame,
            text=f"Task: {current_task['name']}",
            font=('Helvetica', 10),
            wraplength=500
        ).grid(row=1, column=0, sticky="w", pady=2)
        
        # Task description if available
        if current_task['description']:
            ttk.Label(
                task_frame,
                text=f"Description: {current_task['description']}",
                wraplength=500
            ).grid(row=2, column=0, sticky="w", pady=2)
        
        # Task status selection
        status_frame = ttk.Frame(task_frame)
        status_frame.grid(row=3, column=0, sticky="ew", pady=10)
        
        self.task_status = StringVar(value="OK")
        ttk.Label(status_frame, text="Status:").grid(row=0, column=0, padx=5)
        ttk.Radiobutton(
            status_frame,
            text="OK",
            value="OK",
            variable=self.task_status
        ).grid(row=0, column=1, padx=5)
        ttk.Radiobutton(
            status_frame,
            text="Not OK",
            value="NOT_OK", 
            variable=self.task_status
        ).grid(row=0, column=2, padx=5)

    def _create_navigation_frame(self, parent):
        """Create navigation buttons and progress indicator."""
        # Navigation frame at bottom right
        nav_frame = ttk.Frame(parent)
        nav_frame.grid(row=2, column=0, sticky="se", pady=10)
        
        # Previous button
        ttk.Button(
            nav_frame,
            text="Previous",
            command=self._previous_task,
            state="disabled" if self._is_first_task() else "normal"
        ).grid(row=0, column=0, padx=5)
        
        # Next button
        ttk.Button(
            nav_frame,
            text="Next",
            command=self._next_task
        ).grid(row=0, column=1, padx=5)
        
        # Progress indicator
        total_tasks = self._get_total_tasks()
        current_task_num = self._get_current_task_number()
        progress_text = f"Task {current_task_num} of {total_tasks}"
        ttk.Label(nav_frame, text=progress_text).grid(row=0, column=2, padx=20)
        
    def run(self):
        """Start the application."""
        self.root.mainloop()
        
    def _get_total_tasks(self) -> int:
        """Get total number of tasks across all sheets, excluding comment tasks."""
        total = 0
        for sheet in self.data:
            for category in sheet['categories']:
                for task in category['tasks']:
                    if task['name'].lower().strip() != "comments:":
                        total += 1
        return total

    def _get_current_task_number(self) -> int:
        """Get current task number out of total tasks, excluding comment tasks."""
        current = 0
        for s_idx, sheet in enumerate(self.data):
            for c_idx, category in enumerate(sheet['categories']):
                for t_idx, task in enumerate(category['tasks']):
                    if task['name'].lower().strip() != "comments:":
                        current += 1
                    if (s_idx == self.current_sheet and 
                        c_idx == self.current_category and 
                        t_idx == self.current_task):
                        return current
        return current

    def _is_first_task(self) -> bool:
        """Check if currently on first task."""
        return (self.current_sheet == 0 and 
                self.current_category == 0 and 
                self.current_task == 0)

    def _is_last_task(self) -> bool:
        """Check if currently on last task."""
        last_sheet = len(self.data) - 1
        last_category = len(self.data[last_sheet]['categories']) - 1
        last_task = len(self.data[last_sheet]['categories'][last_category]['tasks']) - 1
        return (self.current_sheet == last_sheet and 
                self.current_category == last_category and 
                self.current_task == last_task)

    def _next_task(self):
        """Move to next task."""
        try:
            if self._is_last_task():
                self._complete_checklist()
                return
            self._save_current_task()
            
            while True:
                # Try to move to next task
                if self.current_task < len(self.data[self.current_sheet]['categories'][self.current_category]['tasks']) - 1:
                    self.current_task += 1
                else:
                    # Try next category
                    if self.current_category < len(self.data[self.current_sheet]['categories']) - 1:
                        self.current_category += 1
                        self.current_task = 0
                    else:
                        # Try next sheet
                        if self.current_sheet < len(self.data) - 1:
                            self.current_sheet += 1
                            self.current_category = 0
                            self.current_task = 0
                        else:
                            self._complete_checklist()
                            return
                
                # Check if current task is a comment task
                current_task = self._get_current_task()
                if current_task and current_task['name'].lower().strip() != "comments:":
                    break

            self._show_task_interface()
                
        except Exception as e:
            logger.error(f"Error navigating to next task: {e}")
            logger.debug(f"Current indices - Sheet: {self.current_sheet}, Category: {self.current_category}, Task: {self.current_task}")
            logger.debug(f"Data lengths - Sheets: {len(self.data)}, Categories: {len(self.data[self.current_sheet]['categories'])}, Tasks: {len(self.data[self.current_sheet]['categories'][self.current_category]['tasks'])}")
            messagebox.showerror("Error", "Failed to move to next task")
        
    def _previous_task(self):
        """Move to previous task."""
        if self._is_first_task():
            return
            
        # Move to previous task
        if self.current_task > 0:
            self.current_task -= 1
        else:
            if self.current_category > 0:
                self.current_category -= 1
                self.current_task = len(self.data[self.current_sheet]['categories'][self.current_category]['tasks']) - 1
            else:
                self.current_sheet -= 1
                self.current_category = len(self.data[self.current_sheet]['categories']) - 1
                self.current_task = len(self.data[self.current_sheet]['categories'][self.current_category]['tasks']) - 1
                
        self._show_task_interface()

    def _save_current_task(self):
        """Save the current task's status and comments."""
        try:
            current_task = self._get_current_task()
            if not current_task:
                return
                
            # Skip if task name is "Comments:"
            if current_task['name'].lower().strip() == "comments:":
                return
                
            # Find next empty input cell
            empty_cell = None
            for cell, value in current_task['inputs'].items():
                if value.lower() == "no input":
                    empty_cell = cell
                    break
                    
            if empty_cell:
                # Format initials and date in uppercase
                date_str = f"{self.initials.get().upper()} {self.selected_day.get()} {self.selected_month.get().upper()} {self.selected_year.get()[2:]}"
                current_task['inputs'][empty_cell] = date_str
                
                # Find next empty cell for status
                next_cell = None
                for cell, value in current_task['inputs'].items():
                    if value.lower() == "no input":
                        next_cell = cell
                        break
                        
                if next_cell:
                    current_task['inputs'][next_cell] = self.task_status.get().upper()
                    
        except Exception as e:
            logger.error(f"Error saving task: {e}")
            messagebox.showerror("Error", f"Failed to save task: {str(e)}")

    def _validate_task_indices(self) -> bool:
        """Validate current task indices are within bounds."""
        try:
            if self.current_sheet < 0 or self.current_sheet >= len(self.data):
                return False
                
            current_sheet = self.data[self.current_sheet]
            if ('categories' not in current_sheet or 
                self.current_category < 0 or 
                self.current_category >= len(current_sheet['categories'])):
                return False
                
            current_category = current_sheet['categories'][self.current_category]
            if ('tasks' not in current_category or 
                self.current_task < 0 or 
                self.current_task >= len(current_category['tasks'])):
                return False
                
            return True
            
        except Exception as e:
            logger.error(f"Error validating task indices: {e}")
            return False

    def _complete_checklist(self):
        """Handle checklist completion."""
        try:
            # Save final task
            self._save_current_task()
            
            # Save updated data back to Excel
            self._save_to_excel()
            
            # Close the application
            self.root.destroy()
                
        except Exception as e:
            logger.error(f"Error completing checklist: {e}")
            messagebox.showerror(
                "Error",
                "Failed to complete checklist. Please try again."
            )

    def _clear_previous_month_data(self, workbook):
        """Clear all data from previous month in input cells, excluding rows 1 through 8 and column A."""
        for worksheet in workbook.worksheets:
            # Get all input cell columns from first task's inputs
            input_columns = set()
            for task in self.data[0]['categories'][0]['tasks']:
                for cell_ref in task['inputs'].keys():
                    input_columns.add(cell_ref[0])  # Get column letter
            
            # Clear all cells in input columns, excluding rows 1 through 8 and column A
            for column in input_columns:
                if column == 'A':
                    continue  # Skip column A
                col_idx = openpyxl.utils.column_index_from_string(column)
                for row in range(9, worksheet.max_row + 1):  # Start from row 9
                    if row < 1 or col_idx < 1:
                        continue  # Skip invalid rows or columns
                    cell = worksheet.cell(row=row, column=col_idx)
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = None

    def _preserve_header(self, source_ws, target_ws):
        """
        Preserve rows 1 through 8 and columns A through O from the source worksheet
        and transfer them to the target worksheet, including formatting and merged cells.
        
        Args:
            source_ws: Source worksheet to copy from
            target_ws: Target worksheet to copy to
        """
        for row in range(1, 9):  # Rows 1 through 8
            for col in range(1, 16):  # Columns A through O (1 through 15)
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_ws.cell(row=row, column=col, value=source_cell.value)
                
                # Copy cell styles
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = source_cell.protection.copy()
                    target_cell.alignment = source_cell.alignment.copy()

        # Copy merged cells
        for merge_range in source_ws.merged_cells.ranges:
            if merge_range.min_row <= 8 and merge_range.min_col <= 15:
                target_ws.merge_cells(str(merge_range))

    def _save_to_excel(self):
        """Save data back to Excel file with proper month transition handling."""
        try:
            xlsx_path = self.json_path.parent / f"{self.json_path.stem}.xlsx"
            if not xlsx_path.exists():
                raise FileNotFoundError(f"Original Excel file not found: {xlsx_path}")
                
            # Parse filename
            filename_parts = xlsx_path.stem.split()
            base_name = ' '.join(filename_parts[:-3])
            old_month = filename_parts[-2]
            
            # Format new date
            new_date = f"{self.selected_day.get()} {self.selected_month.get().upper()} {self.selected_year.get()[2:]}"
            new_filename = f"{base_name} {new_date}.xlsx"
            new_path = xlsx_path.parent / new_filename
            
            # Load workbook
            wb = openpyxl.load_workbook(xlsx_path)
            
            # Only clear data if month has changed
            if old_month != self.selected_month.get().upper():
                self._clear_previous_month_data(wb)
                
            # Write new data to worksheets
            for sheet_data in self.data:
                sheet_index = int(sheet_data['name'].split()[-1]) - 1
                if 0 <= sheet_index < len(wb.worksheets):
                    ws = wb.worksheets[sheet_index]
                    if old_month == self.selected_month.get().upper():
                        # Same month - find next empty columns
                        self._write_same_month_data(ws, sheet_data)
                    else:
                        # New month - write to first columns
                        self._write_new_month_data(ws, sheet_data)
                                
            # Save and cleanup
            wb.save(str(new_path))
            self.json_path.unlink()  # Delete JSON file
            
            messagebox.showinfo("Success", f"Excel file saved as: {new_path}")
            
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            messagebox.showerror("Error", f"Failed to save Excel file: {str(e)}")

    def _write_new_month_data(self, worksheet, sheet_data):
        """
        Write new month data to first available cells.
        
        Args:
            worksheet: Worksheet to write to
            sheet_data: Sheet data containing tasks
        """
        try:
            for category in sheet_data['categories']:
                for task in category['tasks']:
                    # Skip comment tasks
                    if task['name'].lower().strip() == "comments:":
                        continue
                        
                    # Get first two empty cells for this task
                    input_cells = sorted(task['inputs'].keys())
                    if len(input_cells) >= 2:
                        # Write to first cell (date and initials)
                        cell = self._get_unmerged_cell(worksheet, input_cells[0])
                        if cell:
                            cell.value = (f"{self.initials.get().upper()} "
                                        f"{self.selected_day.get()} "
                                        f"{self.selected_month.get().upper()} "
                                        f"{self.selected_year.get()[2:]}")
                        
                        # Write to second cell (status)
                        cell = self._get_unmerged_cell(worksheet, input_cells[1])
                        if cell:
                            cell.value = self.task_status.get().upper()
                            
        except Exception as e:
            logger.error(f"Error writing new month data: {e}")
            raise

    def _get_unmerged_cell(self, worksheet, cell_ref):
        """
        Get unmerged cell reference for writing.
        
        Args:
            worksheet: openpyxl Worksheet object
            cell_ref: Cell reference (e.g. 'B10')
            
        Returns:
            Cell object that can be written to
        """
        try:
            # Split cell reference into column and row
            import re
            column_letter = ''.join(filter(str.isalpha, cell_ref))
            row_number = int(''.join(filter(str.isdigit, cell_ref)))
            
            # Convert column letter to index
            from openpyxl.utils import column_index_from_string
            column = column_index_from_string(column_letter)
            
            # Check if cell is in a merged range
            for merge_range in worksheet.merged_cells.ranges:
                if (merge_range.min_row <= row_number <= merge_range.max_row and 
                    merge_range.min_col <= column <= merge_range.max_col):
                    # Return top-left cell of merge range
                    return worksheet.cell(merge_range.min_row, merge_range.min_col)
                    
            # Return requested cell if not merged
            return worksheet.cell(row_number, column)
            
        except Exception as e:
            logger.error(f"Error getting unmerged cell {cell_ref}: {e}")
            return None

    def _update_existing_month_data(self, worksheet, sheet_data):
        """
        Update data in existing month's spreadsheet.
        
        Args:
            worksheet: Worksheet to update
            sheet_data: Sheet data containing tasks
        """
        for category in sheet_data['categories']:
            for task in category['tasks']:
                # Skip comments
                if task['name'].lower().strip() == "comments:":
                    continue
                    
                for cell_ref, value in task['inputs'].items():
                    try:
                        cell = worksheet[cell_ref]
                        if isinstance(cell, openpyxl.cell.cell.MergedCell):
                            continue
                            
                        if value.lower() != "no input":
                            cell.value = value
                            
                    except Exception as e:
                        logger.error(f"Error writing to cell {cell_ref}: {e}")

    def _find_next_empty_columns(self, worksheet, task):
        """Find next two empty columns for the task."""
        input_cells = sorted(task['inputs'].keys())
        empty_columns = []
        
        for cell_ref in input_cells:
            cell = self._get_unmerged_cell(worksheet, cell_ref)
            if cell and not cell.value:
                empty_columns.append(cell_ref)
                if len(empty_columns) == 2:
                    return empty_columns
                    
        return empty_columns if empty_columns else input_cells[:2]

    def _write_same_month_data(self, worksheet, sheet_data):
        """Write data to next empty columns when in same month."""
        try:
            for category in sheet_data['categories']:
                for task in category['tasks']:
                    # Skip comment tasks
                    if task['name'].lower().strip() == "comments:":
                        continue
                        
                    # Find next two empty columns
                    empty_cells = self._find_next_empty_columns(worksheet, task)
                    if len(empty_cells) >= 2:
                        # Write to first empty cell (date and initials)
                        cell = self._get_unmerged_cell(worksheet, empty_cells[0])
                        if cell:
                            cell.value = (f"{self.initials.get().upper()} "
                                        f"{self.selected_day.get()} "
                                        f"{self.selected_month.get().upper()} "
                                        f"{self.selected_year.get()[2:]}")
                        
                        # Write to second empty cell (status)
                        cell = self._get_unmerged_cell(worksheet, empty_cells[1])
                        if cell:
                            cell.value = self.task_status.get().upper()
                            
        except Exception as e:
            logger.error(f"Error writing same month data: {e}")
            raise

    def _load_and_validate_data(self, json_path: Path) -> List[Dict]:
        """Load JSON data and validate structure."""
        try:
            with open(json_path) as f:
                data = json.load(f)
                
            if not isinstance(data, list) or not data:
                raise ValueError("Data must be a non-empty list of sheets")
                
            for sheet in data:
                if 'categories' not in sheet or not isinstance(sheet['categories'], list):
                    raise ValueError(f"Sheet missing valid categories: {sheet}")
                    
                for category in sheet['categories']:
                    if 'tasks' not in category or not isinstance(category['tasks'], list):
                        logger.warning(f"Category missing tasks: {category}")
                        category['tasks'] = []
                        
            return data
            
        except Exception as e:
            logger.error(f"Failed to load task data: {e}")
            messagebox.showerror("Error", f"Failed to load task data: {str(e)}")
            self.root.destroy()
            raise

    def _get_current_task(self) -> Optional[Dict]:
        """
        Safely get current non-comment task.
        
        Returns:
            Dict: Task data if valid task found, None otherwise
        """
        try:
            if not self._validate_indices():
                return None
                
            current_task = (self.data[self.current_sheet]
                           ['categories'][self.current_category]
                           ['tasks'][self.current_task])
            
            return current_task
            
        except Exception as e:
            logger.error(f"Error getting current task: {e}")
            return None

    def _validate_indices(self) -> bool:
        """Validate current task indices."""
        try:
            if not (0 <= self.current_sheet < len(self.data)):
                return False
                
            sheet = self.data[self.current_sheet]
            if not (0 <= self.current_category < len(sheet['categories'])):
                return False
                
            category = sheet['categories'][self.current_category]
            if not (0 <= self.current_task < len(category['tasks'])):
                return False
                
            return True
            
        except Exception as e:
            logger.error(f"Error validating indices: {e}")
            return False

# Usage
if __name__ == "__main__":
    json_path = Path("path/to/processed/template.json")
    app = TaskCompletionUI(json_path)
    app.run()