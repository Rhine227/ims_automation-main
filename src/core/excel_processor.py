"""Excel workbook processing functionality.

This module handles the processing and data extraction from Excel workbooks,
specifically designed for IMS templates.

TODO:
- Add input validation for file paths
- Add progress callbacks for long-running operations
- Add error recovery mechanisms
- Add support for different Excel formats
- Add caching mechanism for processed data
"""

from pathlib import Path
from tkinter.filedialog import askopenfilename
from typing import Dict, List, Set, Optional
import logging
import json
from tkinter import Tk, messagebox
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from ..config.settings import INPUT_HEADERS, CATEGORY_COLOR, DEFAULT_VALUE
from .data_models import Task, Category, SheetData

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Handles Excel file processing and data extraction."""
    
    def __init__(self, file_path: Path):
        """Initialize the Excel processor.
        
        Args:
            file_path: Path to Excel file to process
        """
        self.file_path = file_path
        self.workbook = None
        self._input_cols = None

    def process_workbook(self) -> List[SheetData]:
        """Process entire workbook and return structured data.
        
        Returns:
            List of SheetData objects containing processed worksheet data
            
        Raises:
            Exception: If workbook processing fails
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            return [self._process_worksheet(sheet) 
                   for sheet in self.workbook.worksheets]
        except Exception as e:
            logger.error(f"Error processing workbook: {e}")
            raise

    def _process_worksheet(self, worksheet: Worksheet) -> SheetData:
        """Process a single worksheet and extract structured data."""
        sheet_number = list(self.workbook.worksheets).index(worksheet) + 1
        sheet_data = SheetData(name=f"Sheet {sheet_number}")
        current_category = None
        
        # Find input columns first (excluding column A)
        self._input_cols = self._find_input_columns(worksheet)
        
        for row in worksheet.iter_rows():
            first_cell = row[0]
            if not first_cell.value:
                continue
                
            cell_value = str(first_cell.value).strip()
            
            # Category detection (yellow background and bold)
            if (first_cell.fill.start_color.index == CATEGORY_COLOR and 
                first_cell.font.bold):
                if current_category:
                    sheet_data.categories.append(current_category)
                current_category = Category(name=cell_value)
                
            # Task detection (bold text)
            elif first_cell.font.bold:
                if current_category:
                    task = Task(
                        name=cell_value,
                        inputs=self._get_input_values(row)
                    )
                    current_category.tasks.append(task)
            
            # Description detection (non-bold text)
            elif not first_cell.font.bold and current_category and current_category.tasks:
                current_task = current_category.tasks[-1]
                if current_task.description:
                    current_task.description += " " + cell_value
                else:
                    current_task.description = cell_value

        # Add final category
        if current_category:
            sheet_data.categories.append(current_category)
            
        return sheet_data

    def _find_input_columns(self, worksheet) -> List[int]:
        """Find columns containing input headers (excluding column A)."""
        input_cols = []
        for row in worksheet.iter_rows(max_row=10):
            for cell in row:
                if (cell.value in INPUT_HEADERS and 
                    cell.column > 1):  # Skip column A
                    input_cols.append(cell.column)
        return sorted(set(input_cols))

    def _get_input_values(self, row) -> Dict[str, str]:
        """Extract input values from specified columns (excluding column A)."""
        inputs = {}
        for col in self._input_cols:
            cell = row[col - 1]  # Convert to 0-based index
            coord = f"{openpyxl.utils.get_column_letter(cell.column)}{cell.row}"
            inputs[coord] = str(cell.value) if cell.value else DEFAULT_VALUE
        return inputs

def save_to_json(data: List[SheetData], output_path: Path) -> None:
    """Save the given data to a JSON file."""
    try:
        json_data = [
            {
                "name": sheet.name,
                "categories": [
                    {
                        "name": cat.name,
                        "tasks": [
                            {
                                "name": task.name,
                                "description": task.description,
                                "inputs": task.inputs
                            }
                            for task in cat.tasks
                        ]
                    }
                    for cat in sheet.categories
                ]
            }
            for sheet in data
        ]
        
        with output_path.open('w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=4, ensure_ascii=False)
            
    except Exception as e:
        logger.error(f"Failed to save JSON: {e}")
        raise

def main(template: Optional[str] = None) -> None:
    """Main execution function."""
    root = Tk()
    root.withdraw()

    try:
        if template:
            file_path = Path("IMS_TEMPLATE_COPIES") / template / f"{template}.xlsx"
            if not file_path.exists():
                raise FileNotFoundError(f"Template not found: {file_path}")
        else:
            file_path = Path(askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls")]
            ))
            if not file_path.name:
                logger.info("No file selected")
                return

        processor = ExcelProcessor(file_path)
        data = processor.process_workbook()
        save_to_json(data, file_path.with_suffix('.json'))

    except Exception as e:
        logger.error(f"Error processing file: {e}")
        messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    main()
